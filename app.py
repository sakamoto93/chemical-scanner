from fastapi import FastAPI, File, UploadFile
from fastapi.responses import StreamingResponse, HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import cv2
import numpy as np
from paddleocr import PaddleOCR
import io
from PIL import Image
import re
import time
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pubchempy as pcp
import openpyxl
import os

app = FastAPI()
ocr = PaddleOCR(use_textline_orientation=True, lang='en')

MAX_OCR_DIMENSION = 1200  # OCR処理時間短縮のため、長辺をこのサイズに制限

# 使用するカメラのインデックス（複数カメラ接続時に外付けWebcamを選択するため）
# 環境変数 CAMERA_INDEX で切り替え可能。例: CAMERA_INDEX=1 python app.py
CAMERA_INDEX = int(os.environ.get("CAMERA_INDEX", "0"))
print(f"📷 Using camera index: {CAMERA_INDEX} (環境変数 CAMERA_INDEX で変更可能)")

# リスク対象化合物リストをグローバルにロード
RISK_ASSESSMENT_COMPOUNDS = {}
RISK_ASSESSMENT_METADATA = {}

def load_risk_assessment_list():
    """労働安全衛生法に基づくリスク対象化合物リストを読み込む（正規化済みCSV形式）

    CSVフォーマット: cas_number,compound_name,related_cas,source_sheet
    - related_cas: 同一物質が複数CAS番号を持つ場合、カンマ区切りで全CAS番号を保持
      （元のExcelで「71-23-8, 67-63-0」のように1セルに複数CASがあった行は、
       移行スクリプト scripts/migrate_risk_assessment.py によって
       CAS番号ごとに1行ずつに分割済み。そのためどちらのCAS番号で検索しても
       同じ related_cas を持つレコードがヒットする）

    元の複数シートExcelファイルからこのCSVを生成するには:
      python scripts/migrate_risk_assessment.py <元のxlsxファイル> data/risk_assessment.csv
    """
    global RISK_ASSESSMENT_COMPOUNDS, RISK_ASSESSMENT_METADATA

    possible_paths = [
        "data/risk_assessment.csv",
        os.path.expanduser("~/data/risk_assessment.csv"),
    ]

    risk_list_file = None
    for path in possible_paths:
        if os.path.exists(path):
            risk_list_file = path
            print(f"✅ Found risk assessment file: {path}")
            break
        else:
            print(f"   ⏭️  Not found: {path}")

    if not risk_list_file:
        print("⚠️  Risk assessment file not found in any location:")
        for path in possible_paths:
            print(f"    - {path}")
        print("📋 To enable risk assessment, place 'risk_assessment.csv' in the 'data/' directory")
        print("   (元のExcelファイルから生成する場合は scripts/migrate_risk_assessment.py を使用)")
        return False

    try:
        with open(risk_list_file, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                cas_number = row.get("cas_number", "").strip()
                compound_name = row.get("compound_name", "").strip()
                related_cas_str = row.get("related_cas", "").strip()
                source_sheet = row.get("source_sheet", "").strip()

                if not cas_number or not compound_name:
                    continue

                related_cas = [c.strip() for c in related_cas_str.split(",") if c.strip()] or [cas_number]

                RISK_ASSESSMENT_COMPOUNDS[cas_number] = {
                    "name": compound_name,
                    "related_cas": related_cas,
                    "sheet": source_sheet,
                }

        total_compounds = len(RISK_ASSESSMENT_COMPOUNDS)
        print(f"✅ Loaded {total_compounds} risk assessment compounds")
        if RISK_ASSESSMENT_COMPOUNDS:
            print("   📋 Sample loaded compounds:")
            for i, (cas, info) in enumerate(list(RISK_ASSESSMENT_COMPOUNDS.items())[:5]):
                print(f"      - {cas}: {info['name']}")

        RISK_ASSESSMENT_METADATA["loaded"] = True
        RISK_ASSESSMENT_METADATA["total"] = total_compounds
        return True

    except Exception as e:
        print(f"❌ Error loading risk assessment list: {e}")
        return False

# サーバー起動時にリスト読み込み
print("\n" + "="*60)
print("🔍 Risk Assessment System Initialization")
print("="*60)
load_risk_assessment_list()
if RISK_ASSESSMENT_METADATA.get("loaded"):
    print(f"✅ Risk assessment system ready: {RISK_ASSESSMENT_METADATA.get('total', 0)} compounds loaded")
else:
    print("❌ Risk assessment system failed to load")
print("="*60 + "\n")

def resize_for_ocr(image_array, max_dimension=MAX_OCR_DIMENSION):
    """OCR処理速度向上のため画像をリサイズ（長辺が max_dimension を超える場合のみ縮小）"""
    height, width = image_array.shape[:2]
    longer_side = max(height, width)
    if longer_side <= max_dimension:
        return image_array
    scale = max_dimension / longer_side
    new_size = (int(width * scale), int(height * scale))
    return cv2.resize(image_array, new_size, interpolation=cv2.INTER_AREA)

# Static files
app.mount("/static", StaticFiles(directory="static"), name="static")

def extract_cas_numbers(texts):
    """OCRテキストからすべてのCAS番号を抽出（複数対応）"""
    cas_pattern = r'\d{2,7}-\d{2}-\d'
    cas_numbers = []
    for text_obj in texts:
        text = text_obj.get('text', '') if isinstance(text_obj, dict) else text_obj
        # findall で複数マッチするすべてのCAS番号を取得
        matches = re.findall(cas_pattern, text)
        for match in matches:
            if match not in cas_numbers:  # 重複排除
                cas_numbers.append(match)
    return cas_numbers if cas_numbers else None

def check_risk_assessment(cas_number):
    """CAS番号がリスク対象化合物リストに含まれているか確認

    1つの物質が複数のCAS番号を持つ場合（related_cas）、
    そのうちどのCAS番号で検索してもこの関数は同じレコードをヒットさせる。
    これはロード時（load_risk_assessment_list）に、複数CAS番号を持つ行を
    CAS番号ごとの個別キーとしてすべて登録しているため。
    """
    if not cas_number or not RISK_ASSESSMENT_COMPOUNDS:
        print(f"  [check_risk_assessment] Skipped: cas_number={cas_number}, compounds_loaded={len(RISK_ASSESSMENT_COMPOUNDS)}")
        return None

    if cas_number in RISK_ASSESSMENT_COMPOUNDS:
        compound_info = RISK_ASSESSMENT_COMPOUNDS[cas_number]
        print(f"  [check_risk_assessment] ✅ FOUND: {cas_number} in loaded list")
        return {
            "is_risk_target": True,
            "name": compound_info["name"],
            "sheet": compound_info["sheet"],
            "related_cas": compound_info["related_cas"],
            "regulation": "労働安全衛生法に基づくラベル表示・SDS交付等の義務対象物質"
        }

    print(f"  [check_risk_assessment] ❌ NOT FOUND: {cas_number} (checked {len(RISK_ASSESSMENT_COMPOUNDS)} compounds)")
    # 最初の10個のキーを表示（デバッグ用）
    if len(RISK_ASSESSMENT_COMPOUNDS) > 0:
        first_keys = list(RISK_ASSESSMENT_COMPOUNDS.keys())[:10]
        print(f"    First 10 CAS in loaded list: {first_keys}")
    return None

def validate_cas_checkdigit(cas_number):
    """CAS番号のチェックディジットを検証"""
    try:
        parts = cas_number.split('-')
        if len(parts) != 3:
            return False

        main_part = parts[0] + parts[1]
        check_digit = int(parts[2])

        total = 0
        for i, digit in enumerate(reversed(main_part)):
            total += int(digit) * (i + 1)

        calculated_check = (10 - (total % 10)) % 10
        return check_digit == calculated_check
    except:
        return False

def extract_common_name(compound):
    """synonymsから通称名を抽出（日本語または簡潔な名前を優先）"""
    try:
        if hasattr(compound, 'synonyms') and compound.synonyms:
            # 日本語名またはカタカナ名を探す
            for syn in compound.synonyms:
                syn_str = str(syn)
                # 日本語またはカタカナが含まれているか確認
                if any('぀' <= c <= 'ゟ' or '゠' <= c <= 'ヿ' or '一' <= c <= '鿿' for c in syn_str):
                    return syn_str
            # 見つからない場合は、ノイズを除外しながら最初の有効な同義語を返す
            for syn in compound.synonyms:
                syn_str = str(syn).strip()

                # 除外条件：ノイズとみなされるパターン
                if is_noise_name(syn_str):
                    continue

                # 有効な候補：50文字以下で、IUPAC名と異なる
                if len(syn_str) < 50 and syn_str.lower() != getattr(compound, 'iupac_name', '').lower():
                    return syn_str
    except Exception as e:
        pass
    return ''

def is_noise_name(name):
    """通称名がノイズ（製品情報、カタログ番号など）かどうかを判定"""
    # CAS番号パターン（xx-xx-x）
    if re.search(r'\d{2,7}-\d{2}-\d(?!\w)', name):
        return True
    # EC番号パターン（xxx-xxx-x）
    if re.search(r'\d{3}-\d{3}-\d', name):
        return True
    # グラム数などの単位表記（数字 + 単位）
    if re.search(r'\d+\s*(g|kg|mg|µg|ml|l|mol)\b', name, re.IGNORECASE):
        return True
    # データベースID接頭辞（NSC-, SCHEMBL, DTXSID, MFCD, CHEBI, PUBCHEM など）
    if re.match(r'^(NSC|SCHEMBL|DTXSID|MFCD|CHEBI|PUBCHEM|EC|UNII|EINECS)-', name, re.IGNORECASE):
        return True
    # 純粋に数字だけ（カタログ番号など）
    if re.match(r'^\d+(-\d+)*$', name):
        return True
    # 括弧内のパッケージサイズ情報（括弧が50%以上を占める）
    paren_count = name.count('(') + name.count(')')
    if paren_count > 2 or (len(name) > 20 and paren_count / len(name) > 0.3):
        return True

    return False

def search_pubchem_by_cas(cas_number):
    """CAS番号からPubChemで化合物情報を取得"""
    try:
        # CAS番号で検索（'name'パラメータを使用）
        compounds = pcp.get_compounds(cas_number, 'name')

        if compounds:
            compound = compounds[0]
            return {
                "cas": cas_number,
                "name": getattr(compound, 'iupac_name', 'N/A'),
                "common_name": extract_common_name(compound),
                "formula": getattr(compound, 'molecular_formula', 'N/A'),
                "weight": getattr(compound, 'molecular_weight', 'N/A'),
                "cid": compound.cid,
                "source": "cas"
            }
    except Exception as e:
        pass
    return None

def search_pubchem_by_name(compound_name):
    """化合物名からPubChemで化合物情報を取得"""
    try:
        # 化合物名で検索
        compounds = pcp.get_compounds(compound_name, 'name')

        if compounds:
            compound = compounds[0]
            # CAS番号を取得（複数の属性を試す）
            cas_number = 'N/A'

            # 方法1: iupac_nameから抽出（CAS番号が含まれることがある）
            if hasattr(compound, 'iupac_name'):
                iupac = getattr(compound, 'iupac_name', '')
                cas_match = re.search(r'\d{2,7}-\d{2}-\d', str(iupac))
                if cas_match:
                    cas_number = cas_match.group(0)

            # 方法2: synonymsから抽出
            if cas_number == 'N/A' and hasattr(compound, 'synonyms'):
                try:
                    for syn in compound.synonyms:
                        syn_str = str(syn)
                        cas_match = re.search(r'\d{2,7}-\d{2}-\d', syn_str)
                        if cas_match:
                            cas_number = cas_match.group(0)
                            break
                except:
                    pass

            # 方法3: CIDから詳細情報を取得
            if cas_number == 'N/A' and hasattr(compound, 'cid'):
                try:
                    compound_data = pcp.get_compound(compound.cid, namespace='cid')
                    if hasattr(compound_data, 'iupac_name'):
                        iupac = getattr(compound_data, 'iupac_name', '')
                        cas_match = re.search(r'\d{2,7}-\d{2}-\d', str(iupac))
                        if cas_match:
                            cas_number = cas_match.group(0)
                except:
                    pass

            return {
                "cas": cas_number,
                "name": getattr(compound, 'iupac_name', compound_name),
                "common_name": extract_common_name(compound),
                "formula": getattr(compound, 'molecular_formula', 'N/A'),
                "weight": getattr(compound, 'molecular_weight', 'N/A'),
                "cid": compound.cid,
                "source": "name"
            }
    except Exception as e:
        pass
    return None

def search_compound_by_name_with_risk(compound_name):
    """化合物名でPubChem検索し、取得したCAS番号でリスク判定も行う"""
    compound_info = search_pubchem_by_name(compound_name)
    risk_assessment = None
    if compound_info and compound_info.get('cas') and compound_info['cas'] != 'N/A':
        risk_assessment = check_risk_assessment(compound_info['cas'])
    return compound_info, risk_assessment

def get_camera_frame():
    cap = cv2.VideoCapture(CAMERA_INDEX)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # JPEG エンコード
        ret, buffer = cv2.imencode('.jpg', frame)
        frame_bytes = buffer.tobytes()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n'
               b'Content-Length: ' + f"{len(frame_bytes)}".encode() + b'\r\n\r\n'
               + frame_bytes + b'\r\n')

@app.get("/", response_class=HTMLResponse)
async def get_index():
    with open("templates/index.html", "r", encoding="utf-8") as f:
        return f.read()

@app.get("/video_feed")
async def video_feed():
    return StreamingResponse(
        get_camera_frame(),
        media_type="multipart/x-mixed-replace; boundary=frame"
    )
@app.get("/capture")
async def capture_frame():
    cap = cv2.VideoCapture(CAMERA_INDEX)
    ret, frame = cap.read()
    cap.release()
    
    if not ret:
        return JSONResponse({"error": "Failed to capture frame"}, status_code=400)
    
    ret, buffer = cv2.imencode('.jpg', frame)
    frame_bytes = buffer.tobytes()
    
    return StreamingResponse(
        iter([frame_bytes]),
        media_type="image/jpeg"
    )
@app.post("/ocr")
async def perform_ocr(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))
        image_array = np.array(image)
        original_shape = image_array.shape[:2]
        image_array = resize_for_ocr(image_array)

        ocr_start = time.time()
        result = ocr.ocr(image_array)
        ocr_elapsed = time.time() - ocr_start
        print(f"[OCR Benchmark] Image resized {original_shape} -> {image_array.shape[:2]}, "
              f"ocr.ocr() took {ocr_elapsed:.2f}s")

        texts = []
        if result and isinstance(result, list) and len(result) > 0:
            result_dict = result[0]
            if isinstance(result_dict, dict):
                rec_texts = result_dict.get('rec_texts', [])
                rec_scores = result_dict.get('rec_scores', [])
                for text, score in zip(rec_texts, rec_scores):
                    texts.append({
                        "text": text,
                        "confidence": float(score)
                    })

        # CAS番号を抽出（複数対応）
        cas_numbers = extract_cas_numbers(texts)
        cas_number = cas_numbers[0] if cas_numbers else None  # 最初のCAS番号を使用
        compound_info = None
        risk_assessment = None

        print(f"[/ocr] Extracted CAS numbers: {cas_numbers}")

        if cas_numbers:
            # すべてのCAS番号に対してリスク判定を実行
            for cas in cas_numbers:
                risk_result = check_risk_assessment(cas)
                if risk_result:
                    # 最初のリスク対象化合物を使用
                    risk_assessment = risk_result
                    cas_number = cas  # リスク対象が見つかった場合、そのCAS番号を優先
                    print(f"[/ocr] Found risk target: CAS={cas}, {risk_result}")
                    break

            # リスク対象でない場合は、最初のCAS番号を使用
            if not risk_assessment and cas_numbers:
                cas_number = cas_numbers[0]

            # Phase 1: CAS番号で検索
            if cas_number:
                compound_info = search_pubchem_by_cas(cas_number)
                if not risk_assessment:
                    # Phase 1で見つけた場合、改めてリスク判定
                    risk_assessment = check_risk_assessment(cas_number)
                    print(f"[/ocr] Risk assessment result: {risk_assessment}")
        else:
            # Phase 2: 化合物名で検索（信頼度が高いテキストから順に試す）
            # 信頼度でソート（降順）
            sorted_texts = sorted(texts, key=lambda x: x['confidence'], reverse=True)

            for text_obj in sorted_texts:
                compound_name = text_obj['text']
                confidence = text_obj['confidence']

                # 信頼度が一定以上のテキストのみ検索
                if confidence >= 0.85:
                    result, name_risk_assessment = search_compound_by_name_with_risk(compound_name)
                    if result:
                        compound_info = result
                        risk_assessment = name_risk_assessment
                        print(f"[/ocr] Found compound from name '{compound_name}': {result.get('name')}, CAS={result.get('cas')}")
                        print(f"[/ocr] Risk assessment for {result.get('cas')}: {risk_assessment}")
                        break

        response_data = {
            "texts": texts,
            "cas_numbers": cas_numbers,  # すべてのCAS番号
            "cas_number": cas_number,    # 優先されるCAS番号（最初またはリスク対象）
            "compound_info": compound_info,
            "risk_assessment": risk_assessment
        }
        print(f"[/ocr] Final response: all_cas_numbers={cas_numbers}, primary_cas={cas_number}, has_risk_assessment={risk_assessment is not None}")
        return JSONResponse(response_data)
    except Exception as e:
        import traceback
        return JSONResponse({"error": str(e), "traceback": traceback.format_exc()}, status_code=400)

@app.post("/search_by_name")
async def search_by_name(data: dict):
    """試薬名を手入力してPubChem検索し、CAS番号・リスク判定を行う

    OCRでCAS番号が読み取れない、あるいは化合物名がPubChemの検索に
    うまくヒットしない試薬瓶に対して、ユーザーが試薬名を直接入力して
    検索できるようにするためのエンドポイント。
    """
    try:
        compound_name = data.get("compound_name", "").strip()
        if not compound_name:
            return JSONResponse({"error": "化合物名を入力してください"}, status_code=400)

        print(f"[/search_by_name] Searching for: {compound_name}")
        compound_info, risk_assessment = search_compound_by_name_with_risk(compound_name)

        if not compound_info:
            print(f"[/search_by_name] Not found: {compound_name}")
            return JSONResponse({
                "compound_info": None,
                "risk_assessment": None,
                "error": f"「{compound_name}」はPubChemで見つかりませんでした"
            })

        print(f"[/search_by_name] Found: {compound_info.get('name')}, CAS={compound_info.get('cas')}, risk={risk_assessment is not None}")
        return JSONResponse({
            "compound_info": compound_info,
            "risk_assessment": risk_assessment
        })
    except Exception as e:
        import traceback
        return JSONResponse({"error": str(e), "traceback": traceback.format_exc()}, status_code=400)

@app.post("/export/excel")
async def export_excel(data: dict):
    """リスト内容をExcelファイルとして返却"""
    try:
        compounds = data.get("compounds", [])

        # Workbookを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "Compounds"

        # ヘッダーを作成
        headers = ["CAS番号", "化合物名", "検出方法", "分子式", "分子量", "通称名", "リスク対象"]
        ws.append(headers)

        # ヘッダーのスタイルを設定（青背景、白文字、太字）
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # データを追加
        for compound in compounds:
            cas = compound.get("cas", "N/A")
            name = compound.get("name", "N/A")
            source = compound.get("source", "不明")
            source_text = "CAS番号から取得" if source == "cas" else "化合物名から取得"
            formula = compound.get("formula", "N/A")
            weight = compound.get("weight", "N/A")
            common_name = compound.get("commonName", "")
            risk_info = compound.get("riskAssessment", {})
            risk_text = "対象" if (risk_info and risk_info.get("is_risk_target")) else "-"

            ws.append([cas, name, source_text, formula, weight, common_name, risk_text])

        # 列幅を自動調整
        column_widths = [15, 50, 15, 15, 12, 20, 15]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width

        # ファイルをメモリに保存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # タイムスタンプ付きファイル名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"chemical_list_{timestamp}.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        return JSONResponse({"error": str(e), "traceback": traceback.format_exc()}, status_code=400)

@app.post("/export/csv")
async def export_csv(data: dict):
    """リスト内容をCSVとして返却"""
    try:
        compounds = data.get("compounds", [])

        # CSVをメモリに作成
        output = io.StringIO()
        writer = csv.writer(output)

        # ヘッダーを書き込み
        headers = ["CAS番号", "化合物名", "検出方法", "分子式", "分子量", "通称名", "リスク対象"]
        writer.writerow(headers)

        # データを書き込み
        for compound in compounds:
            cas = compound.get("cas", "N/A")
            name = compound.get("name", "N/A")
            source = compound.get("source", "不明")
            source_text = "CAS番号から取得" if source == "cas" else "化合物名から取得"
            formula = compound.get("formula", "N/A")
            weight = compound.get("weight", "N/A")
            common_name = compound.get("commonName", "")
            risk_info = compound.get("riskAssessment", {})
            risk_text = "対象" if (risk_info and risk_info.get("is_risk_target")) else "-"

            writer.writerow([cas, name, source_text, formula, weight, common_name, risk_text])

        csv_content = output.getvalue()

        # タイムスタンプ付きファイル名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"chemical_list_{timestamp}.csv"

        return StreamingResponse(
            iter([csv_content.encode('utf-8-sig')]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        return JSONResponse({"error": str(e), "traceback": traceback.format_exc()}, status_code=400)

if __name__ == "__main__":
    import uvicorn
    import os

    # HTTPS対応（証明書ファイルが存在する場合）
    ssl_keyfile = "key.pem" if os.path.exists("key.pem") else None
    ssl_certfile = "cert.pem" if os.path.exists("cert.pem") else None

    if ssl_keyfile and ssl_certfile:
        print("🔒 Starting with HTTPS (port 8443)")
        print("   Access from MacBook: https://localhost:8443")
        print("   Access from iPhone:  https://172.20.10.x:8443")
        uvicorn.run(
            app,
            host="0.0.0.0",
            port=8443,
            ssl_keyfile=ssl_keyfile,
            ssl_certfile=ssl_certfile
        )
    else:
        print("🚀 Starting with HTTP (port 8000)")
        print("   Access from MacBook: http://localhost:8000")
        print("   ⚠️  Note: mediaDevices requires HTTPS on non-localhost")
        uvicorn.run(app, host="0.0.0.0", port=8000)
