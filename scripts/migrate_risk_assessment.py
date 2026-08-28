#!/usr/bin/env python3
"""
リスク対象化合物リストの移行スクリプト

複数シート・複数フォーマットの労働安全衛生法対象物質Excelファイルを、
単一のフラットなCSVファイル（data/risk_assessment.csv）に変換する。

【変換ルール】
1. 全シートを走査し、ヘッダー行（「名称」「化合物名」「CAS」を含む行）を自動検出
2. データ行から CAS番号・化合物名を抽出
3. 1セルに複数CAS番号がカンマ等で列挙されている場合、CAS番号ごとに行を分割
   （例: "71-23-8, 67-63-0" → 2行に分割。どちらのCAS番号でも検索ヒットするように
    related_cas 列に元の複数CAS番号をすべて記録する）
4. 同一CAS番号が複数シートに重複登場する場合は、最初に見つかったものを採用し、
   出典シート名を source_sheet に記録する

【使い方】
  python scripts/migrate_risk_assessment.py <入力Excelファイル> [出力CSVファイル]

  例:
  python scripts/migrate_risk_assessment.py data/risk_assessment.xlsx data/risk_assessment.csv
"""
import sys
import re
import csv
import openpyxl


def find_header_and_columns(ws, max_scan_rows=50):
    """シート内のヘッダー行と、名称列・CAS列の位置を検出する

    注意: 「英語名称」列は「名称」を部分文字列として含むため、
    単純な部分一致だと日本語名列より優先されてしまうバグがあった。
    そのため、まず完全一致（「名称」「化合物名」など）を優先的に探し、
    見つからない場合のみ部分一致にフォールバックする。
    また「英語名称」「English」等を含む列は名称列の候補から除外する。
    """
    exact_name_labels = {'名称', '化合物名', 'Name', '名前'}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx > max_scan_rows:
            break

        if row and any(cell and ('名称' in str(cell) or 'CAS' in str(cell) or '化合物' in str(cell)) for cell in row):
            name_col = None
            name_col_fallback = None
            cas_col = None

            for col_idx, cell_val in enumerate(row):
                if not cell_val:
                    continue
                cell_str_raw = str(cell_val).strip()
                cell_str_lower = cell_str_raw.lower()

                # 英語名称列は除外
                is_english_name_col = '英語' in cell_str_raw or 'english' in cell_str_lower

                if cell_str_raw in exact_name_labels and not is_english_name_col:
                    name_col = col_idx
                elif not is_english_name_col and any(term in cell_str_raw for term in ['名称', '化合物名', 'Name', '名前']):
                    if name_col_fallback is None:
                        name_col_fallback = col_idx

                if 'cas' in cell_str_lower:
                    cas_col = col_idx

            if name_col is None:
                name_col = name_col_fallback

            if name_col is not None and cas_col is not None:
                return row_idx, name_col, cas_col

    return None, None, None


def split_cas_numbers(cas_cell_value):
    """1セル内の複数CAS番号（カンマ・読点・スラッシュ区切りなど）をすべて抽出"""
    if not cas_cell_value:
        return []
    cas_str = str(cas_cell_value)
    cas_pattern = r'\d{2,7}-\d{2}-\d'
    return re.findall(cas_pattern, cas_str)


def migrate(input_path, output_path):
    print(f"📂 Reading: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=True)

    # cas_number -> {compound_name, related_cas(list), source_sheet}
    compounds = {}
    total_rows_scanned = 0
    total_multi_cas_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, name_col, cas_col = find_header_and_columns(ws)

        if header_row is None:
            print(f"  ⚠️  Sheet '{sheet_name}': header row not found, skipping")
            continue

        print(f"  📋 Sheet '{sheet_name}': header_row={header_row}, name_col={name_col}, cas_col={cas_col}")

        sheet_count = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx <= header_row:
                continue
            if not row or all(cell is None for cell in row):
                continue

            cas_val = row[cas_col] if cas_col < len(row) else None
            name_val = row[name_col] if name_col < len(row) else None

            if not cas_val or not name_val:
                continue

            cas_numbers = split_cas_numbers(cas_val)
            if not cas_numbers:
                continue

            total_rows_scanned += 1
            if len(cas_numbers) > 1:
                total_multi_cas_rows += 1

            compound_name = str(name_val).strip()

            # CAS番号ごとに行を作る。related_cas には同グループの全CAS番号を持たせる
            for cas in cas_numbers:
                if cas not in compounds:
                    compounds[cas] = {
                        "compound_name": compound_name,
                        "related_cas": cas_numbers,
                        "source_sheet": sheet_name,
                    }
                    sheet_count += 1

        print(f"     → {sheet_count} new compounds registered from this sheet")

    print(f"\n✅ Total unique CAS numbers: {len(compounds)}")
    print(f"   Rows with multiple CAS numbers (comma-separated etc.): {total_multi_cas_rows}")

    # CSV出力
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["cas_number", "compound_name", "related_cas", "source_sheet"])
        for cas, info in sorted(compounds.items()):
            related = ",".join(info["related_cas"])
            writer.writerow([cas, info["compound_name"], related, info["source_sheet"]])

    print(f"\n💾 Saved: {output_path}")
    print(f"   ({len(compounds)} rows written)")

    # サンプル表示
    print("\n📋 Sample rows:")
    for i, (cas, info) in enumerate(sorted(compounds.items())):
        if i >= 5:
            break
        related_note = f" (関連CAS: {', '.join(info['related_cas'])})" if len(info['related_cas']) > 1 else ""
        print(f"   {cas}: {info['compound_name']}{related_note}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python scripts/migrate_risk_assessment.py <入力Excelファイル> [出力CSVファイル]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "data/risk_assessment.csv"

    migrate(input_path, output_path)
